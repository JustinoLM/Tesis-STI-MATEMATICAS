"""
Servicio de importación del Excel completo (inverso del exportador de
admin_exports.py). Reconstruye estudiantes, diagnóstico/post-test, sesiones,
niveles, medallas y tienda a partir de las 8 hojas del export.

Modo "sandbox seguro": todo lo importado queda aislado en una organización
elegida explícitamente por el admin — nunca se tocan organizaciones
existentes ni se sobreescriben estudiantes reales (el código de estudiante
es único por organización, así que un mismo código en una org nueva jamás
choca con el de la org real).
"""

import ast
import re
from datetime import datetime
from typing import Optional

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl.worksheet.worksheet import Worksheet

from app.core.security import get_password_hash
from app.models.organization import Organizacion
from app.models.user import Estudiante, TipoUsuario
from app.models.adaptive import (
    PerfilEstudiante,
    PruebaDiagnostica,
    ResultadoPostTest,
    SesionPractica,
    EstadoDiagnostico,
    EstadoSesion,
)
from app.models.gamification import (
    Medalla,
    EstudianteMedalla,
    Desbloqueable,
    EstudianteDesbloqueable,
    TransaccionPuntos,
    TipoTransaccion,
)
from app.schemas.admin_import import ImportResumen, ConteoHoja


# ============================================
# Utilidades de parseo (columnas planas -> tipos Python)
# ============================================

def _int_or_none(v) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _parse_fecha(v) -> Optional[datetime]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(str(v).strip())
    except ValueError:
        return None


def _es_si(v) -> bool:
    return str(v).strip().lower() in ("sí", "si") if v else False


def _parse_duracion(v) -> Optional[int]:
    """'14m 7s' -> 847 segundos."""
    if not v:
        return None
    m = re.match(r"(?:(\d+)\s*m)?\s*(?:(\d+)\s*s)?", str(v).strip())
    if not m or (not m.group(1) and not m.group(2)):
        return None
    return int(m.group(1) or 0) * 60 + int(m.group(2) or 0)


_SIMBOLO_A_NOMBRE = {"+": "suma", "-": "resta", "×": "mult", "x": "mult", "*": "mult", "÷": "div", "/": "div"}


def _parse_operaciones(v) -> Optional[dict]:
    """"{'+': 5, '-': 10}" o "{'suma': 5}" -> {'suma': 5, 'resta': 10}."""
    if not v:
        return None
    try:
        parsed = ast.literal_eval(str(v))
    except (ValueError, SyntaxError):
        return None
    if not isinstance(parsed, dict):
        return None
    return {_SIMBOLO_A_NOMBRE.get(k, k): val for k, val in parsed.items()}


def _nivel_desde_correctos(correctos: int) -> int:
    return 3 if correctos == 2 else (2 if correctos == 1 else 1)


def _booleans_desde_correctos(correctos: int) -> tuple[bool, bool]:
    """El Excel solo trae el conteo (0/1/2), no cuál pregunta específica
    fue correcta. Convención: con 1 correcto se asume que fue la de nivel 1
    (no afecta el nivel calculado, que depende solo del conteo)."""
    if correctos >= 2:
        return True, True
    if correctos == 1:
        return True, False
    return False, False


class AdminImportService:
    """Servicio de importación de Excel completo a una organización sandbox."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.advertencias: list[str] = []

    # ============================================
    # Orquestación
    # ============================================

    async def importar_excel(self, wb, organizacion_nombre: str) -> ImportResumen:
        self.advertencias = []
        org, creada = await self._get_or_create_organizacion(organizacion_nombre)
        hojas: list[ConteoHoja] = []
        codigo_to_id: dict[str, int] = {}

        if "Estudiantes" in wb.sheetnames:
            codigo_to_id, conteo = await self._importar_estudiantes(wb["Estudiantes"], org.id)
            hojas.append(conteo)
        else:
            self.advertencias.append("No se encontró la hoja 'Estudiantes' — no se puede importar nada más sin ella.")

        if codigo_to_id:
            if "Diagnóstico" in wb.sheetnames:
                hojas.append(await self._importar_diagnostico(wb["Diagnóstico"], codigo_to_id))
            if "Sesiones" in wb.sheetnames:
                hojas.append(await self._importar_sesiones(wb["Sesiones"], codigo_to_id))
            if "Niveles Actuales" in wb.sheetnames:
                hojas.append(await self._importar_niveles_actuales(wb["Niveles Actuales"], codigo_to_id))
            if "Medallas" in wb.sheetnames:
                hojas.append(await self._importar_medallas(wb["Medallas"], codigo_to_id))
            if "Tienda" in wb.sheetnames:
                hojas.append(await self._importar_tienda(wb["Tienda"], codigo_to_id))

        return ImportResumen(
            organizacion_id=org.id,
            organizacion_nombre=org.nombre,
            organizacion_creada=creada,
            hojas=hojas,
            advertencias=self.advertencias,
        )

    # ============================================
    # Organización
    # ============================================

    async def _get_or_create_organizacion(self, nombre: str) -> tuple[Organizacion, bool]:
        result = await self.db.execute(select(Organizacion).where(Organizacion.nombre == nombre))
        org = result.scalar_one_or_none()
        if org:
            return org, False

        base_slug = re.sub(r"[^A-Z0-9]+", "-", nombre.upper()).strip("-")[:40] or "SANDBOX"
        codigo = base_slug
        sufijo = 0
        while True:
            result = await self.db.execute(select(Organizacion).where(Organizacion.codigo == codigo))
            if not result.scalar_one_or_none():
                break
            sufijo += 1
            codigo = f"{base_slug}-{sufijo}"

        org = Organizacion(
            nombre=nombre,
            codigo=codigo,
            descripcion="Organización sandbox creada por importación de Excel — datos de prueba.",
            pais="Panamá",
        )
        self.db.add(org)
        await self.db.commit()
        await self.db.refresh(org)
        return org, True

    # ============================================
    # Estudiantes
    # ============================================

    async def _get_estudiante_by_codigo_org(self, codigo: str, org_id: int) -> Optional[Estudiante]:
        result = await self.db.execute(
            select(Estudiante).where(Estudiante.codigo_estudiante == codigo, Estudiante.organizacion_id == org_id)
        )
        return result.scalar_one_or_none()

    async def _importar_estudiantes(self, ws: Worksheet, org_id: int) -> tuple[dict[str, int], ConteoHoja]:
        conteo = ConteoHoja(hoja="Estudiantes")
        codigo_to_id: dict[str, int] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            conteo.procesados += 1
            codigo = str(row[0]).strip()

            existente = await self._get_estudiante_by_codigo_org(codigo, org_id)
            if existente:
                codigo_to_id[codigo] = existente.id
                conteo.omitidos += 1
                continue

            genero = str(row[2] or "masculino").strip().lower()
            if genero not in ("masculino", "femenino"):
                genero = "masculino"
            password_plain = str(row[10]) if row[10] else "cambiar123"

            estudiante = Estudiante(
                codigo_estudiante=codigo,
                password_hash=get_password_hash(password_plain),
                password_plain=password_plain,
                tipo_usuario=TipoUsuario.ESTUDIANTE,
                nombre_completo=row[1] or codigo,
                genero=genero,
                organizacion_id=org_id,
                grado_academico=row[3],
                edad=_int_or_none(row[4]),
                puntos_totales=_int_or_none(row[6]) or 0,
                activo=True,
            )
            self.db.add(estudiante)
            await self.db.flush()  # necesitamos el id sin cerrar la transacción

            self.db.add(PerfilEstudiante(estudiante_id=estudiante.id, nivel_actual=1))

            codigo_to_id[codigo] = estudiante.id
            conteo.creados += 1

        await self.db.commit()
        return codigo_to_id, conteo

    # ============================================
    # Diagnóstico (pre-test / post-test)
    # ============================================

    async def _get_or_create_perfil(self, estudiante_id: int) -> PerfilEstudiante:
        result = await self.db.execute(
            select(PerfilEstudiante).where(PerfilEstudiante.estudiante_id == estudiante_id)
        )
        perfil = result.scalar_one_or_none()
        if not perfil:
            perfil = PerfilEstudiante(estudiante_id=estudiante_id, nivel_actual=1)
            self.db.add(perfil)
            await self.db.flush()
        return perfil

    async def _importar_diagnostico(self, ws: Worksheet, codigo_to_id: dict[str, int]) -> ConteoHoja:
        conteo = ConteoHoja(hoja="Diagnóstico")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            codigo = str(row[0]).strip()
            estudiante_id = codigo_to_id.get(codigo)
            if not estudiante_id:
                self.advertencias.append(f"Diagnóstico: estudiante '{codigo}' no está en la hoja Estudiantes, fila omitida")
                continue
            conteo.procesados += 1
            perfil = await self._get_or_create_perfil(estudiante_id)

            if _es_si(row[4]):
                existe = await self.db.execute(
                    select(PruebaDiagnostica).where(PruebaDiagnostica.estudiante_id == estudiante_id)
                )
                if existe.scalar_one_or_none():
                    conteo.omitidos += 1
                else:
                    pre_suma = _int_or_none(row[6]) or 0
                    pre_resta = _int_or_none(row[7]) or 0
                    pre_mult = _int_or_none(row[8]) or 0
                    pre_div = _int_or_none(row[9]) or 0
                    pre_total = _int_or_none(row[10])
                    if pre_total is None:
                        pre_total = pre_suma + pre_resta + pre_mult + pre_div
                    pre_fecha = _parse_fecha(row[5])

                    s1, s2 = _booleans_desde_correctos(pre_suma)
                    r1, r2 = _booleans_desde_correctos(pre_resta)
                    m1, m2 = _booleans_desde_correctos(pre_mult)
                    d1, d2 = _booleans_desde_correctos(pre_div)

                    if pre_total == 8:
                        nivel_suma = nivel_resta = nivel_mult = nivel_div = 2
                        nivel_actual = 4
                        perfecto = True
                    else:
                        nivel_suma = _int_or_none(row[11]) or _nivel_desde_correctos(pre_suma)
                        nivel_resta = _int_or_none(row[12]) or _nivel_desde_correctos(pre_resta)
                        nivel_mult = _int_or_none(row[13]) or _nivel_desde_correctos(pre_mult)
                        nivel_div = _int_or_none(row[14]) or _nivel_desde_correctos(pre_div)
                        nivel_actual = max(nivel_suma, nivel_resta, nivel_mult, nivel_div)
                        perfecto = False

                    self.db.add(PruebaDiagnostica(
                        estudiante_id=estudiante_id,
                        estado=EstadoDiagnostico.COMPLETADO,
                        fecha_inicio=pre_fecha,
                        fecha_fin=pre_fecha,
                        suma_nivel1_correcto=s1, suma_nivel2_correcto=s2,
                        resta_nivel1_correcto=r1, resta_nivel2_correcto=r2,
                        mult_nivel1_correcto=m1, mult_nivel2_correcto=m2,
                        div_nivel1_correcto=d1, div_nivel2_correcto=d2,
                        nivel_suma_asignado=nivel_suma,
                        nivel_resta_asignado=nivel_resta,
                        nivel_mult_asignado=nivel_mult,
                        nivel_div_asignado=nivel_div,
                        nivel_actual_asignado=nivel_actual,
                        perfecto=perfecto,
                    ))

                    perfil.nivel_suma = nivel_suma
                    perfil.nivel_resta = nivel_resta
                    perfil.nivel_multiplicacion = nivel_mult
                    perfil.nivel_division = nivel_div
                    perfil.nivel_actual = nivel_actual
                    perfil.diagnostico_completado = True
                    perfil.fecha_diagnostico = pre_fecha
                    conteo.creados += 1

            if _es_si(row[15]):
                existe_post = await self.db.execute(
                    select(ResultadoPostTest).where(ResultadoPostTest.estudiante_id == estudiante_id)
                )
                if existe_post.scalar_one_or_none():
                    conteo.omitidos += 1
                else:
                    post_suma = _int_or_none(row[17]) or 0
                    post_resta = _int_or_none(row[18]) or 0
                    post_mult = _int_or_none(row[19]) or 0
                    post_div = _int_or_none(row[20]) or 0
                    post_total = _int_or_none(row[21])
                    if post_total is None:
                        post_total = post_suma + post_resta + post_mult + post_div
                    post_fecha = _parse_fecha(row[16])

                    s1, s2 = _booleans_desde_correctos(post_suma)
                    r1, r2 = _booleans_desde_correctos(post_resta)
                    m1, m2 = _booleans_desde_correctos(post_mult)
                    d1, d2 = _booleans_desde_correctos(post_div)

                    self.db.add(ResultadoPostTest(
                        estudiante_id=estudiante_id,
                        completado=True,
                        fecha_inicio=post_fecha,
                        fecha_fin=post_fecha,
                        suma_nivel1_correcto=s1, suma_nivel2_correcto=s2,
                        resta_nivel1_correcto=r1, resta_nivel2_correcto=r2,
                        mult_nivel1_correcto=m1, mult_nivel2_correcto=m2,
                        div_nivel1_correcto=d1, div_nivel2_correcto=d2,
                        nivel_suma_evaluado=_nivel_desde_correctos(post_suma),
                        nivel_resta_evaluado=_nivel_desde_correctos(post_resta),
                        nivel_mult_evaluado=_nivel_desde_correctos(post_mult),
                        nivel_div_evaluado=_nivel_desde_correctos(post_div),
                        pre_nivel_suma=perfil.nivel_suma,
                        pre_nivel_resta=perfil.nivel_resta,
                        pre_nivel_mult=perfil.nivel_multiplicacion,
                        pre_nivel_div=perfil.nivel_division,
                        pre_nivel_actual=perfil.nivel_actual,
                        total_correctos=post_total,
                        perfecto=(post_total == 8),
                    ))
                    conteo.creados += 1

        await self.db.commit()
        return conteo

    # ============================================
    # Sesiones
    # ============================================

    async def _importar_sesiones(self, ws: Worksheet, codigo_to_id: dict[str, int]) -> ConteoHoja:
        conteo = ConteoHoja(hoja="Sesiones")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            codigo = str(row[1]).strip()
            estudiante_id = codigo_to_id.get(codigo)
            if not estudiante_id:
                self.advertencias.append(f"Sesiones: estudiante '{codigo}' no está en la hoja Estudiantes, fila omitida")
                continue
            conteo.procesados += 1

            self.db.add(SesionPractica(
                estudiante_id=estudiante_id,
                perfil_id=estudiante_id,
                estado=EstadoSesion.COMPLETADA,
                fecha_inicio=_parse_fecha(row[4]),
                fecha_fin=_parse_fecha(row[5]),
                nivel_actual_inicio=_int_or_none(row[11]),
                cantidad_problemas=_int_or_none(row[7]) or 0,
                operaciones_incluidas=_parse_operaciones(row[14]),
                problemas_correctos=_int_or_none(row[8]) or 0,
                problemas_incorrectos=_int_or_none(row[9]) or 0,
                tiempo_total_segundos=_parse_duracion(row[6]),
                es_practica_perfecta=_es_si(row[13]),
                puntos_ganados=_int_or_none(row[12]) or 0,
            ))
            conteo.creados += 1

        await self.db.commit()
        return conteo

    # ============================================
    # Niveles Actuales (estado final autoritativo del perfil)
    # ============================================

    async def _importar_niveles_actuales(self, ws: Worksheet, codigo_to_id: dict[str, int]) -> ConteoHoja:
        conteo = ConteoHoja(hoja="Niveles Actuales")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            codigo = str(row[0]).strip()
            estudiante_id = codigo_to_id.get(codigo)
            if not estudiante_id:
                self.advertencias.append(f"Niveles Actuales: estudiante '{codigo}' no está en la hoja Estudiantes, fila omitida")
                continue
            conteo.procesados += 1
            perfil = await self._get_or_create_perfil(estudiante_id)

            nivel_suma = _int_or_none(row[3])
            nivel_resta = _int_or_none(row[4])
            nivel_mult = _int_or_none(row[5])
            nivel_div = _int_or_none(row[6])
            nivel_general = _int_or_none(row[7])
            total_sesiones = _int_or_none(row[8])

            if nivel_suma is not None:
                perfil.nivel_suma = nivel_suma
            if nivel_resta is not None:
                perfil.nivel_resta = nivel_resta
            if nivel_mult is not None:
                perfil.nivel_multiplicacion = nivel_mult
            if nivel_div is not None:
                perfil.nivel_division = nivel_div
            if nivel_general is not None:
                perfil.nivel_actual = nivel_general
            if total_sesiones is not None:
                perfil.total_sesiones = total_sesiones

            precision_raw = row[9]
            if precision_raw:
                try:
                    perfil.precision_ultimos_15 = float(str(precision_raw).replace("%", "").strip()) / 100
                except ValueError:
                    pass

            conteo.creados += 1

        await self.db.commit()
        return conteo

    # ============================================
    # Medallas
    # ============================================

    async def _estudiante_tiene_medalla(self, estudiante_id: int, medalla_id: int) -> bool:
        # Conteo (no scalar_one_or_none) — el Excel fuente puede traer la misma
        # medalla repetida para un estudiante (ej. "100 Club" otorgada varias
        # veces por un bug del generador de datos), y no hay unique constraint
        # en la tabla real que lo impida.
        result = await self.db.execute(
            select(func.count(EstudianteMedalla.id)).where(
                EstudianteMedalla.estudiante_id == estudiante_id,
                EstudianteMedalla.medalla_id == medalla_id,
            )
        )
        return (result.scalar() or 0) > 0

    async def _importar_medallas(self, ws: Worksheet, codigo_to_id: dict[str, int]) -> ConteoHoja:
        conteo = ConteoHoja(hoja="Medallas")
        medalla_cache: dict[str, Optional[int]] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            codigo = str(row[0]).strip()
            estudiante_id = codigo_to_id.get(codigo)
            if not estudiante_id:
                self.advertencias.append(f"Medallas: estudiante '{codigo}' no está en la hoja Estudiantes, fila omitida")
                continue
            conteo.procesados += 1

            nombre_medalla = str(row[3] or "").strip()
            if not nombre_medalla:
                continue

            if nombre_medalla not in medalla_cache:
                result = await self.db.execute(select(Medalla).where(Medalla.nombre == nombre_medalla).limit(1))
                medalla = result.scalar_one_or_none()
                medalla_cache[nombre_medalla] = medalla.id if medalla else None
                if not medalla:
                    self.advertencias.append(f"Medallas: '{nombre_medalla}' no existe en el catálogo de medallas, se omite")

            medalla_id = medalla_cache[nombre_medalla]
            if medalla_id is None:
                conteo.omitidos += 1
                continue

            if await self._estudiante_tiene_medalla(estudiante_id, medalla_id):
                conteo.omitidos += 1
                continue

            self.db.add(EstudianteMedalla(
                estudiante_id=estudiante_id,
                medalla_id=medalla_id,
                fecha_obtencion=_parse_fecha(row[6]) or datetime.utcnow(),
                notificada=True,
            ))
            conteo.creados += 1

        await self.db.commit()
        return conteo

    # ============================================
    # Tienda
    # ============================================

    async def _estudiante_posee_item(self, estudiante_id: int, desbloqueable_id: int) -> bool:
        result = await self.db.execute(
            select(func.count(EstudianteDesbloqueable.id)).where(
                EstudianteDesbloqueable.estudiante_id == estudiante_id,
                EstudianteDesbloqueable.desbloqueable_id == desbloqueable_id,
            )
        )
        return (result.scalar() or 0) > 0

    async def _importar_tienda(self, ws: Worksheet, codigo_to_id: dict[str, int]) -> ConteoHoja:
        conteo = ConteoHoja(hoja="Tienda")
        item_cache: dict[str, Optional[int]] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            codigo = str(row[0]).strip()
            estudiante_id = codigo_to_id.get(codigo)
            if not estudiante_id:
                self.advertencias.append(f"Tienda: estudiante '{codigo}' no está en la hoja Estudiantes, fila omitida")
                continue
            conteo.procesados += 1

            item_raw = str(row[3] or "").strip()
            puntos_gastados = _int_or_none(row[4]) or 0
            saldo_resultante = _int_or_none(row[5]) or 0
            fecha = _parse_fecha(row[6]) or datetime.utcnow()

            if item_raw.startswith("Pista"):
                # Las pistas no son items del catálogo de desbloqueables —
                # solo se registra el gasto de puntos.
                self.db.add(TransaccionPuntos(
                    estudiante_id=estudiante_id,
                    tipo=TipoTransaccion.GASTO,
                    cantidad=-abs(puntos_gastados),
                    concepto=item_raw,
                    saldo_resultante=saldo_resultante,
                    fecha=fecha,
                ))
                conteo.creados += 1
                continue

            nombre_item = re.sub(r"^Compra:\s*", "", item_raw).strip()
            if nombre_item not in item_cache:
                result = await self.db.execute(select(Desbloqueable).where(Desbloqueable.nombre == nombre_item).limit(1))
                item = result.scalar_one_or_none()
                item_cache[nombre_item] = item.id if item else None
                if not item:
                    self.advertencias.append(f"Tienda: item '{nombre_item}' no existe en el catálogo, se registra el gasto sin desbloqueable asociado")

            desbloqueable_id = item_cache[nombre_item]

            self.db.add(TransaccionPuntos(
                estudiante_id=estudiante_id,
                tipo=TipoTransaccion.GASTO,
                cantidad=-abs(puntos_gastados),
                concepto=item_raw,
                desbloqueable_id=desbloqueable_id,
                saldo_resultante=saldo_resultante,
                fecha=fecha,
            ))

            if desbloqueable_id is not None and not await self._estudiante_posee_item(estudiante_id, desbloqueable_id):
                self.db.add(EstudianteDesbloqueable(
                    estudiante_id=estudiante_id,
                    desbloqueable_id=desbloqueable_id,
                    fecha_compra=fecha,
                    puntos_gastados=puntos_gastados,
                ))
            conteo.creados += 1

        await self.db.commit()
        return conteo
